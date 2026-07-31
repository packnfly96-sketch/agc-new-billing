"""Backend API tests for SD ENTERPRISES Courier Billing (iteration 2 - with JWT auth)."""
import io
import os
import base64
from datetime import date

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip()
                break
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "packnfly96@gmail.com"
ADMIN_PASSWORD = "SDEnterprises@2026"

PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8/5+hHgAHggJ/PchI7wAAAABJRU5ErkJggg=="
)


def _fy():
    d = date.today()
    if d.month >= 4:
        s, e = d.year, d.year + 1
    else:
        s, e = d.year - 1, d.year
    return f"{s}-{str(e)[-2:]}"


# ---------------- Fixtures ----------------
@pytest.fixture(scope="session")
def token():
    """Login once and reuse token."""
    r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    data = r.json()
    assert "access_token" in data
    return data["access_token"]


@pytest.fixture(scope="session")
def client(token):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


@pytest.fixture
def unauth_client():
    """Function-scoped so cookies don't leak between auth tests."""
    return requests.Session()


# =====================================================================
# AUTH
# =====================================================================
class TestAuth:
    def test_health_public(self, unauth_client):
        r = unauth_client.get(f"{API}/")
        assert r.status_code == 200
        assert r.json().get("ok") is True

    def test_login_success(self, unauth_client):
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
        assert r.status_code == 200
        d = r.json()
        assert d["access_token"] and d["token_type"] == "bearer"
        assert d["user"]["email"] == ADMIN_EMAIL
        assert d["user"]["role"] == "admin"
        # cookie set
        assert "access_token" in r.cookies

    def test_login_wrong_password(self, unauth_client):
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": "wrong-pw"})
        assert r.status_code == 401
        assert "Invalid" in r.json().get("detail", "")

    def test_me_requires_token(self, unauth_client):
        r = unauth_client.get(f"{API}/auth/me")
        assert r.status_code == 401

    def test_me_with_token(self, client):
        r = client.get(f"{API}/auth/me")
        assert r.status_code == 200
        assert r.json()["email"] == ADMIN_EMAIL

    def test_protected_endpoints_require_auth(self, unauth_client):
        for path in ["/company", "/customers", "/partners", "/invoices", "/reports/summary"]:
            r = unauth_client.get(f"{API}{path}")
            assert r.status_code == 401, f"{path} should require auth, got {r.status_code}"

    def test_protected_endpoints_ok_with_auth(self, client):
        for path in ["/company", "/customers", "/partners", "/invoices", "/reports/summary"]:
            r = client.get(f"{API}{path}")
            assert r.status_code == 200, f"{path} failed: {r.status_code} {r.text}"

    def test_brute_force_lockout(self):
        """5 wrong attempts locks out (verified against localhost since the public
        ingress rotates client IP across requests, defeating the request.client.host
        based counter — see backend_issues in test report)."""
        LOCAL = "http://localhost:8001/api"
        fake_email = f"brute-{os.getpid()}@example.com"
        for i in range(5):
            r = requests.post(f"{LOCAL}/auth/login",
                              json={"email": fake_email, "password": "x"})
            assert r.status_code == 401
        r = requests.post(f"{LOCAL}/auth/login",
                          json={"email": fake_email, "password": "x"})
        assert r.status_code == 429, f"Expected lockout, got {r.status_code}: {r.text}"

    def test_change_password_flow(self, client, unauth_client):
        new_pw = "TempPass@12345"
        # wrong current
        r = client.post(f"{API}/auth/change-password",
                        json={"current_password": "wrong", "new_password": new_pw})
        assert r.status_code == 400
        # too short
        r = client.post(f"{API}/auth/change-password",
                        json={"current_password": ADMIN_PASSWORD, "new_password": "abc"})
        assert r.status_code == 400
        # correct
        r = client.post(f"{API}/auth/change-password",
                        json={"current_password": ADMIN_PASSWORD, "new_password": new_pw})
        assert r.status_code == 200
        # new pw works
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": new_pw})
        assert r.status_code == 200
        new_token = r.json()["access_token"]
        # restore original (using new token as bearer)
        r = requests.post(f"{API}/auth/change-password",
                          headers={"Authorization": f"Bearer {new_token}"},
                          json={"current_password": new_pw, "new_password": ADMIN_PASSWORD})
        assert r.status_code == 200, f"restore failed: {r.text}"
        # verify original works again
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
        assert r.status_code == 200

    def test_forgot_and_reset_password_flow(self, unauth_client):
        r = unauth_client.post(f"{API}/auth/forgot-password", json={"email": ADMIN_EMAIL})
        assert r.status_code == 200
        reset_token = r.json().get("reset_token")
        assert reset_token, "reset_token missing in response"

        temp_pw = "ResetPass@2026"
        r = unauth_client.post(f"{API}/auth/reset-password",
                               json={"token": reset_token, "new_password": temp_pw})
        assert r.status_code == 200
        # login with new
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": temp_pw})
        assert r.status_code == 200
        tok = r.json()["access_token"]
        # restore
        r = requests.post(f"{API}/auth/change-password",
                          headers={"Authorization": f"Bearer {tok}"},
                          json={"current_password": temp_pw, "new_password": ADMIN_PASSWORD})
        assert r.status_code == 200

    def test_reset_password_invalid_token(self, unauth_client):
        r = unauth_client.post(f"{API}/auth/reset-password",
                               json={"token": "bogus", "new_password": "SomePass@2026"})
        assert r.status_code == 400

    def test_logout(self, unauth_client):
        # login gets cookie, logout should clear
        r = unauth_client.post(f"{API}/auth/login",
                               json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
        assert r.status_code == 200
        r = unauth_client.post(f"{API}/auth/logout")
        assert r.status_code == 200
        assert r.json().get("ok") is True


# =====================================================================
# COMPANY
# =====================================================================
class TestCompany:
    def test_get(self, client):
        r = client.get(f"{API}/company")
        assert r.status_code == 200
        assert "name" in r.json()

    def test_update(self, client):
        payload = {
            "name": "SD ENTERPRISES", "address": "123 Test Rd",
            "state": "Maharashtra", "state_code": "27",
            "gstin": "27ABCDE1234F1Z5", "pan": "ABCDE1234F",
            "invoice_prefix": "SDE", "default_tax_rate": 18.0,
        }
        r = client.put(f"{API}/company", json=payload)
        assert r.status_code == 200
        assert r.json()["gstin"] == "27ABCDE1234F1Z5"

    def test_upload_and_get_logo(self, client):
        files = {"file": ("logo.png", PNG_BYTES, "image/png")}
        r = client.post(f"{API}/company/assets/logo", files=files)
        assert r.status_code == 200
        assert r.json()["logo"]["mime"] == "image/png"
        r = client.get(f"{API}/company/assets/logo/file")
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("image/")


# =====================================================================
# CUSTOMERS + PARTNERS
# =====================================================================
@pytest.fixture(scope="session")
def customer(client):
    r = client.post(f"{API}/customers", json={
        "name": "TEST_CUST_Iter2", "state": "Maharashtra", "state_code": "27",
        "gstin": "27AAAPL1234C1Z1", "city": "Mumbai",
    })
    assert r.status_code == 200
    return r.json()


@pytest.fixture(scope="session")
def partner(client):
    r = client.post(f"{API}/partners", json={"name": "TEST_PARTNER_Iter2", "code": "DTDC"})
    assert r.status_code == 200
    return r.json()


class TestCustomerPartner:
    def test_customer_crud(self, client, customer):
        cid = customer["id"]
        r = client.get(f"{API}/customers/{cid}")
        assert r.status_code == 200
        r = client.put(f"{API}/customers/{cid}",
                       json={"name": "TEST_CUST_Iter2_U", "state_code": "27"})
        assert r.json()["name"] == "TEST_CUST_Iter2_U"

    def test_partner_list(self, client, partner):
        r = client.get(f"{API}/partners")
        assert r.status_code == 200
        assert any(p["id"] == partner["id"] for p in r.json())


# =====================================================================
# INVOICES
# =====================================================================
class TestInvoices:
    def test_create_invoice_cgst(self, client, customer, partner):
        r = client.post(f"{API}/invoices", json={
            "customer_id": customer["id"],
            "items": [{"docket_no": "D1", "partner_id": partner["id"],
                       "amount": 500.0, "pieces": 1, "weight": 1}],
            "gst_type": "cgst_sgst", "tax_rate": 18.0,
        })
        assert r.status_code == 200
        inv = r.json()
        assert inv["subtotal"] == 500.0 and inv["total"] == 590.0
        assert inv["cgst"] == 45.0 and inv["sgst"] == 45.0
        assert inv["invoice_number"].startswith("SDE/")
        pytest.shared_invoice_id = inv["id"]

    def test_duplicate_invoice(self, client):
        src_id = pytest.shared_invoice_id
        # fetch source
        src = client.get(f"{API}/invoices/{src_id}").json()
        r = client.post(f"{API}/invoices/{src_id}/duplicate")
        assert r.status_code == 200
        dup = r.json()
        assert dup["id"] != src["id"]
        assert dup["invoice_number"] != src["invoice_number"]
        assert dup["customer_id"] == src["customer_id"]
        assert dup["total"] == src["total"]
        assert dup["status"] == "draft"
        assert dup["payment_status"] == "unpaid"
        # sequence increments
        s_seq = int(src["invoice_number"].split("/")[-1])
        d_seq = int(dup["invoice_number"].split("/")[-1])
        assert d_seq > s_seq
        # original unchanged
        src2 = client.get(f"{API}/invoices/{src_id}").json()
        assert src2["invoice_number"] == src["invoice_number"]

    def test_duplicate_not_found(self, client):
        r = client.post(f"{API}/invoices/nonexistent-id/duplicate")
        assert r.status_code == 404

    def test_invoice_pdf(self, client):
        r = client.get(f"{API}/invoices/{pytest.shared_invoice_id}/pdf")
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"
        assert len(r.content) > 4000  # new blue design is non-trivial


# =====================================================================
# MONTHLY EXCEL
# =====================================================================
class TestMonthlyExcel:
    def test_download_xlsx(self, client):
        r = client.get(f"{API}/reports/monthly-excel", params={"year": 2026, "month": 7})
        assert r.status_code == 200
        assert r.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "SDE_Monthly_2026-07.xlsx" in r.headers.get("content-disposition", "")
        assert len(r.content) > 1000
        # Verify openable & has expected structure
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # header row (row 4)
        headers = [ws.cell(row=4, column=c).value for c in range(1, 11)]
        assert "Invoice #" in headers and "Grand Total" in headers
        # TOTAL row present somewhere in col D
        col_d = [ws.cell(row=r_, column=4).value for r_ in range(5, ws.max_row + 1)]
        assert "TOTAL" in col_d

    def test_invalid_month(self, client):
        r = client.get(f"{API}/reports/monthly-excel", params={"year": 2026, "month": 13})
        assert r.status_code == 400

    def test_excel_requires_auth(self, unauth_client):
        r = unauth_client.get(f"{API}/reports/monthly-excel", params={"year": 2026, "month": 7})
        assert r.status_code == 401


# =====================================================================
# REPORTS
# =====================================================================
class TestReports:
    def test_summary(self, client):
        r = client.get(f"{API}/reports/summary")
        assert r.status_code == 200
        for k in ("totals", "gst", "top_customers", "top_partners", "monthly"):
            assert k in r.json()
