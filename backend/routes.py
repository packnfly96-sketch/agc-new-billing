"""API routes for the courier billing app (all require auth)."""
import uuid
import io
from datetime import datetime, timezone, date
from calendar import monthrange
from typing import List, Optional
from pathlib import Path

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from PIL import Image as PILImage, UnidentifiedImageError

from db import (
    db, UPLOAD_ASSETS, MAX_ASSET_SIZE, ALLOWED_MIME, EXT_BY_MIME,
    COMPANY_SINGLETON_ID,
)
from models import (
    Company, CompanyUpdate, CompanyAsset,
    Customer, CustomerCreate,
    CourierPartner, CourierPartnerCreate,
    Invoice, InvoiceCreate, InvoiceItem,
    now_iso,
)
from pdf_service import build_invoice_pdf
from auth import get_current_user

router = APIRouter(prefix="/api", dependencies=[Depends(get_current_user)])


# --------------------- Helpers ---------------------
ASSET_FIELDS = {"logo", "signature", "stamp"}


async def get_or_create_company() -> Company:
    doc = await db.company.find_one({"_id": COMPANY_SINGLETON_ID})
    if not doc:
        c = Company(name="SD ENTERPRISES")
        payload = c.model_dump()
        payload["_id"] = COMPANY_SINGLETON_ID
        await db.company.insert_one(payload)
        return c
    doc.pop("_id", None)
    return Company(**doc)


def _current_fiscal_year(d: date) -> str:
    # Indian FY starts April
    if d.month >= 4:
        start, end = d.year, d.year + 1
    else:
        start, end = d.year - 1, d.year
    return f"{start}-{str(end)[-2:]}"


async def _next_invoice_number(prefix: str, fiscal_year: str) -> str:
    from pymongo import ReturnDocument
    counter_key = f"invoice::{prefix}::{fiscal_year}"
    doc = await db.counters.find_one_and_update(
        {"_id": counter_key},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = doc["seq"] if doc else 1
    return f"{prefix}/{fiscal_year}/{seq:04d}"


def _round(n: float) -> float:
    return float(f"{n:.2f}")


def _compute_totals(items: List[InvoiceItem], gst_type: str, tax_rate: float) -> dict:
    subtotal = _round(sum(i.amount for i in items))
    cgst = sgst = igst = 0.0
    if gst_type == "cgst_sgst":
        half = tax_rate / 2
        cgst = _round(subtotal * half / 100)
        sgst = _round(subtotal * half / 100)
    elif gst_type == "igst":
        igst = _round(subtotal * tax_rate / 100)
    total_tax = _round(cgst + sgst + igst)
    raw_total = subtotal + total_tax
    total = round(raw_total)
    round_off = _round(total - raw_total)
    return {
        "subtotal": subtotal,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "total_tax": total_tax,
        "round_off": round_off,
        "total": _round(total),
    }


async def _hydrate_partner_names(items: List[InvoiceItem]) -> List[InvoiceItem]:
    # Fill partner_name from partner_id when missing
    ids = [i.partner_id for i in items if i.partner_id and not i.partner_name]
    if not ids:
        return items
    docs = await db.partners.find({"id": {"$in": list(set(ids))}}, {"_id": 0}).to_list(500)
    name_by_id = {d["id"]: d["name"] for d in docs}
    for it in items:
        if it.partner_id and not it.partner_name:
            it.partner_name = name_by_id.get(it.partner_id, "")
    return items


# --------------------- Company ---------------------
@router.get("/company")
async def api_get_company():
    return (await get_or_create_company()).model_dump()


@router.put("/company")
async def api_update_company(payload: CompanyUpdate):
    updates = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if updates:
        updates["updated_at"] = now_iso()
        await db.company.update_one({"_id": COMPANY_SINGLETON_ID}, {"$set": updates}, upsert=True)
    return (await get_or_create_company()).model_dump()


@router.post("/company/assets/{asset_type}")
async def api_upload_asset(asset_type: str, file: UploadFile = File(...)):
    if asset_type not in ASSET_FIELDS:
        raise HTTPException(400, f"Unknown asset type '{asset_type}'. Use logo, signature, or stamp.")
    if file.content_type not in ALLOWED_MIME:
        raise HTTPException(400, "Unsupported file type. Allowed: JPG, PNG, SVG.")
    contents = await file.read()
    if not contents:
        raise HTTPException(400, "Empty file uploaded.")
    if len(contents) > MAX_ASSET_SIZE:
        raise HTTPException(400, "File is too large. Maximum size is 2 MB.")

    # For raster images, verify decodable bytes with Pillow to prevent corrupt files
    # from breaking PDF generation later. SVG is text-based and skipped.
    if file.content_type != "image/svg+xml":
        try:
            with PILImage.open(io.BytesIO(contents)) as im:
                im.verify()
        except (UnidentifiedImageError, OSError, ValueError):
            raise HTTPException(400, "The uploaded file is not a valid image.")

    company = await get_or_create_company()
    prev: Optional[CompanyAsset] = getattr(company, asset_type)
    if prev:
        old = UPLOAD_ASSETS / prev.filename
        if old.exists():
            try:
                old.unlink()
            except OSError:
                pass

    ext = EXT_BY_MIME[file.content_type]
    filename = f"{asset_type}-{uuid.uuid4().hex}{ext}"
    (UPLOAD_ASSETS / filename).write_bytes(contents)

    await db.company.update_one(
        {"_id": COMPANY_SINGLETON_ID},
        {"$set": {
            asset_type: {"filename": filename, "mime": file.content_type},
            "updated_at": now_iso(),
        }},
        upsert=True,
    )
    return (await get_or_create_company()).model_dump()


@router.delete("/company/assets/{asset_type}")
async def api_remove_asset(asset_type: str):
    if asset_type not in ASSET_FIELDS:
        raise HTTPException(400, f"Unknown asset type '{asset_type}'.")
    company = await get_or_create_company()
    prev: Optional[CompanyAsset] = getattr(company, asset_type)
    if prev:
        old = UPLOAD_ASSETS / prev.filename
        if old.exists():
            try:
                old.unlink()
            except OSError:
                pass
    await db.company.update_one(
        {"_id": COMPANY_SINGLETON_ID},
        {"$set": {asset_type: None, "updated_at": now_iso()}},
    )
    return (await get_or_create_company()).model_dump()


@router.get("/company/assets/{asset_type}/file")
async def api_get_asset_file(asset_type: str):
    if asset_type not in ASSET_FIELDS:
        raise HTTPException(400, f"Unknown asset type '{asset_type}'.")
    company = await get_or_create_company()
    asset: Optional[CompanyAsset] = getattr(company, asset_type)
    if not asset:
        raise HTTPException(404, f"No {asset_type} uploaded.")
    path = UPLOAD_ASSETS / asset.filename
    if not path.exists():
        raise HTTPException(404, "File missing on disk.")
    return FileResponse(str(path), media_type=asset.mime or "application/octet-stream")


# --------------------- Customers ---------------------
@router.get("/customers")
async def api_list_customers():
    docs = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return docs


@router.post("/customers")
async def api_create_customer(payload: CustomerCreate):
    if not payload.name.strip():
        raise HTTPException(400, "Customer name is required.")
    c = Customer(**payload.model_dump())
    await db.customers.insert_one(c.model_dump())
    return c.model_dump()


@router.get("/customers/{customer_id}")
async def api_get_customer(customer_id: str):
    doc = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Customer not found")
    return doc


@router.put("/customers/{customer_id}")
async def api_update_customer(customer_id: str, payload: CustomerCreate):
    existing = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Customer not found")
    merged = {**existing, **payload.model_dump()}
    await db.customers.update_one({"id": customer_id}, {"$set": merged})
    return merged


@router.delete("/customers/{customer_id}")
async def api_delete_customer(customer_id: str):
    res = await db.customers.delete_one({"id": customer_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Customer not found")
    return {"ok": True}


# --------------------- Courier Partners ---------------------
@router.get("/partners")
async def api_list_partners():
    docs = await db.partners.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return docs


@router.post("/partners")
async def api_create_partner(payload: CourierPartnerCreate):
    if not payload.name.strip():
        raise HTTPException(400, "Partner name is required.")
    p = CourierPartner(**payload.model_dump())
    await db.partners.insert_one(p.model_dump())
    return p.model_dump()


@router.get("/partners/{partner_id}")
async def api_get_partner(partner_id: str):
    doc = await db.partners.find_one({"id": partner_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Partner not found")
    return doc


@router.put("/partners/{partner_id}")
async def api_update_partner(partner_id: str, payload: CourierPartnerCreate):
    existing = await db.partners.find_one({"id": partner_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Partner not found")
    merged = {**existing, **payload.model_dump()}
    await db.partners.update_one({"id": partner_id}, {"$set": merged})
    return merged


@router.delete("/partners/{partner_id}")
async def api_delete_partner(partner_id: str):
    res = await db.partners.delete_one({"id": partner_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Partner not found")
    return {"ok": True}


# --------------------- Invoices ---------------------
@router.get("/invoices")
async def api_list_invoices():
    docs = await db.invoices.find({}, {"_id": 0}).sort("invoice_date", -1).to_list(5000)
    return docs


@router.get("/invoices/next-number")
async def api_next_number(prefix: Optional[str] = None):
    """Preview the next invoice number without incrementing (for form display)."""
    company = await get_or_create_company()
    p = prefix or company.invoice_prefix or "SDE"
    fy = _current_fiscal_year(date.today())
    counter_key = f"invoice::{p}::{fy}"
    doc = await db.counters.find_one({"_id": counter_key})
    next_seq = (doc["seq"] if doc else 0) + 1
    return {"next_number": f"{p}/{fy}/{next_seq:04d}", "fiscal_year": fy, "prefix": p}


@router.post("/invoices")
async def api_create_invoice(payload: InvoiceCreate):
    company = await get_or_create_company()
    customer = await db.customers.find_one({"id": payload.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(400, "Selected customer does not exist.")

    inv_date = payload.invoice_date or date.today().isoformat()
    fy = _current_fiscal_year(date.fromisoformat(inv_date))
    prefix = company.invoice_prefix or "SDE"
    number = await _next_invoice_number(prefix, fy)

    items = await _hydrate_partner_names(payload.items)
    totals = _compute_totals(items, payload.gst_type, payload.tax_rate)

    inv = Invoice(
        invoice_number=number,
        fiscal_year=fy,
        invoice_date=inv_date,
        due_date=payload.due_date,
        customer_id=customer["id"],
        customer_name=customer["name"],
        customer_address=customer.get("address", ""),
        customer_city=customer.get("city", ""),
        customer_state=customer.get("state", ""),
        customer_state_code=customer.get("state_code", ""),
        customer_pincode=customer.get("pincode", ""),
        customer_gstin=customer.get("gstin", ""),
        customer_phone=customer.get("phone", ""),
        customer_email=customer.get("email", ""),
        items=items,
        gst_type=payload.gst_type,
        tax_rate=payload.tax_rate,
        notes=payload.notes,
        terms=payload.terms or company.default_terms,
        status=payload.status,
        payment_status=payload.payment_status,
        **totals,
    )
    await db.invoices.insert_one(inv.model_dump())
    return inv.model_dump()


@router.get("/invoices/{invoice_id}")
async def api_get_invoice(invoice_id: str):
    doc = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Invoice not found")
    return doc


@router.put("/invoices/{invoice_id}")
async def api_update_invoice(invoice_id: str, payload: InvoiceCreate):
    existing = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Invoice not found")
    customer = await db.customers.find_one({"id": payload.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(400, "Selected customer does not exist.")

    items = await _hydrate_partner_names(payload.items)
    totals = _compute_totals(items, payload.gst_type, payload.tax_rate)

    updated = {
        **existing,
        "invoice_date": payload.invoice_date or existing["invoice_date"],
        "due_date": payload.due_date,
        "customer_id": customer["id"],
        "customer_name": customer["name"],
        "customer_address": customer.get("address", ""),
        "customer_city": customer.get("city", ""),
        "customer_state": customer.get("state", ""),
        "customer_state_code": customer.get("state_code", ""),
        "customer_pincode": customer.get("pincode", ""),
        "customer_gstin": customer.get("gstin", ""),
        "customer_phone": customer.get("phone", ""),
        "customer_email": customer.get("email", ""),
        "items": [i.model_dump() for i in items],
        "gst_type": payload.gst_type,
        "tax_rate": payload.tax_rate,
        "notes": payload.notes,
        "terms": payload.terms,
        "status": payload.status,
        "payment_status": payload.payment_status,
        "updated_at": now_iso(),
        **totals,
    }
    await db.invoices.update_one({"id": invoice_id}, {"$set": updated})
    return updated


@router.delete("/invoices/{invoice_id}")
async def api_delete_invoice(invoice_id: str):
    res = await db.invoices.delete_one({"id": invoice_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Invoice not found")
    return {"ok": True}


@router.get("/invoices/{invoice_id}/pdf")
async def api_invoice_pdf(invoice_id: str):
    doc = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Invoice not found")
    company = await get_or_create_company()
    inv = Invoice(**doc)
    pdf_bytes = build_invoice_pdf(inv, company)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{inv.invoice_number.replace("/", "-")}.pdf"'},
    )


# --------------------- Reports & Utilities ---------------------
class ReportRange(BaseModel):
    start: Optional[str] = None
    end: Optional[str] = None


# Public health check — no auth (attached to a separate router at import time)
public_router = APIRouter(prefix="/api")


@public_router.get("/")
async def api_root():
    return {"ok": True, "service": "SD ENTERPRISES Courier Billing API"}


# --------------------- Duplicate Invoice ---------------------
@router.post("/invoices/{invoice_id}/duplicate")
async def api_duplicate_invoice(invoice_id: str):
    src = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not src:
        raise HTTPException(404, "Invoice not found")
    company = await get_or_create_company()

    inv_date = date.today().isoformat()
    fy = _current_fiscal_year(date.fromisoformat(inv_date))
    prefix = company.invoice_prefix or "SDE"
    number = await _next_invoice_number(prefix, fy)

    new_inv = Invoice(
        invoice_number=number,
        fiscal_year=fy,
        invoice_date=inv_date,
        due_date=src.get("due_date", ""),
        customer_id=src["customer_id"],
        customer_name=src["customer_name"],
        customer_address=src.get("customer_address", ""),
        customer_city=src.get("customer_city", ""),
        customer_state=src.get("customer_state", ""),
        customer_state_code=src.get("customer_state_code", ""),
        customer_pincode=src.get("customer_pincode", ""),
        customer_gstin=src.get("customer_gstin", ""),
        customer_phone=src.get("customer_phone", ""),
        customer_email=src.get("customer_email", ""),
        items=[InvoiceItem(**i) for i in src.get("items", [])],
        gst_type=src.get("gst_type", "none"),
        tax_rate=src.get("tax_rate", 18),
        notes=src.get("notes", ""),
        terms=src.get("terms", ""),
        status="draft",
        payment_status="unpaid",
        subtotal=src.get("subtotal", 0),
        cgst=src.get("cgst", 0),
        sgst=src.get("sgst", 0),
        igst=src.get("igst", 0),
        total_tax=src.get("total_tax", 0),
        round_off=src.get("round_off", 0),
        total=src.get("total", 0),
    )
    await db.invoices.insert_one(new_inv.model_dump())
    return new_inv.model_dump()


# --------------------- Monthly Excel Export ---------------------
@router.get("/reports/monthly-excel")
async def api_monthly_excel(year: int, month: int):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if month < 1 or month > 12:
        raise HTTPException(400, "Invalid month")
    last_day = monthrange(year, month)[1]
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year:04d}-{month:02d}-{last_day:02d}"

    invoices = await db.invoices.find(
        {"invoice_date": {"$gte": start, "$lte": end}}, {"_id": 0}
    ).sort("invoice_date", 1).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # Title
    ws.merge_cells("A1:J1")
    company = await get_or_create_company()
    title = f"{company.name or 'SD ENTERPRISES'} — Monthly Sales Report"
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    month_name = date(year, month, 1).strftime("%B %Y")
    ws["A2"] = f"Period: {month_name}  ·  {start} to {end}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Invoice #", "Date", "Customer", "Courier Partner", "PCS", "Weight (kg)",
               "Amount", "GST", "Round Off", "Grand Total"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(border_style="thin", color="0F172A"))

    row = header_row + 1
    tot_pcs = tot_wt = tot_amt = tot_gst = tot_round = tot_grand = 0.0
    for inv in invoices:
        items = inv.get("items", []) or []
        partners = ", ".join(sorted({(it.get("partner_name") or "").strip() for it in items if it.get("partner_name")})) or "—"
        pcs = sum(int(it.get("pieces", 0) or 0) for it in items)
        wt = sum(float(it.get("weight", 0) or 0) for it in items)
        amount = float(inv.get("subtotal", 0) or 0)
        gst_amt = float(inv.get("total_tax", 0) or 0)
        round_off = float(inv.get("round_off", 0) or 0)
        grand = float(inv.get("total", 0) or 0)

        values = [inv.get("invoice_number", ""), inv.get("invoice_date", ""), inv.get("customer_name", ""),
                  partners, pcs, wt, amount, gst_amt, round_off, grand]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name="Calibri", size=10)
            if col >= 5:
                cell.alignment = Alignment(horizontal="right")
                if col >= 7:
                    cell.number_format = "#,##0.00"
        row += 1
        tot_pcs += pcs
        tot_wt += wt
        tot_amt += amount
        tot_gst += gst_amt
        tot_round += round_off
        tot_grand += grand

    # Totals row
    totals = ["", "", "", "TOTAL", tot_pcs, tot_wt, tot_amt, tot_gst, tot_round, tot_grand]
    for col, v in enumerate(totals, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="right" if col >= 5 else "center")
        if col >= 7:
            cell.number_format = "#,##0.00"

    # Column widths
    widths = [20, 12, 32, 24, 8, 12, 14, 14, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SDE_Monthly_{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------- Reports ---------------------
@router.get("/reports/summary")
async def api_reports_summary(start: Optional[str] = None, end: Optional[str] = None):
    q: dict = {}
    if start or end:
        q["invoice_date"] = {}
        if start:
            q["invoice_date"]["$gte"] = start
        if end:
            q["invoice_date"]["$lte"] = end
    invoices = await db.invoices.find(q, {"_id": 0}).to_list(10000)

    total_invoices = len(invoices)
    gross = sum(i.get("total", 0) for i in invoices)
    tax = sum(i.get("total_tax", 0) for i in invoices)
    taxable = sum(i.get("subtotal", 0) for i in invoices)
    paid = sum(i.get("total", 0) for i in invoices if i.get("payment_status") == "paid")
    unpaid = gross - paid

    # by customer
    by_customer: dict = {}
    for inv in invoices:
        key = inv.get("customer_name", "—")
        by_customer[key] = by_customer.get(key, 0) + inv.get("total", 0)
    top_customers = sorted(by_customer.items(), key=lambda x: -x[1])[:10]

    # by partner (aggregate items across all invoices)
    by_partner: dict = {}
    for inv in invoices:
        for it in inv.get("items", []) or []:
            key = it.get("partner_name") or "Unassigned"
            by_partner[key] = by_partner.get(key, 0) + it.get("amount", 0)
    top_partners = sorted(by_partner.items(), key=lambda x: -x[1])[:10]

    # monthly
    by_month: dict = {}
    for inv in invoices:
        d = inv.get("invoice_date", "")
        month = d[:7] if d else "unknown"
        by_month[month] = by_month.get(month, 0) + inv.get("total", 0)
    monthly = sorted(by_month.items())

    # GST summary
    cgst_total = sum(i.get("cgst", 0) for i in invoices)
    sgst_total = sum(i.get("sgst", 0) for i in invoices)
    igst_total = sum(i.get("igst", 0) for i in invoices)

    return {
        "totals": {
            "invoices": total_invoices,
            "gross": round(gross, 2),
            "taxable": round(taxable, 2),
            "tax": round(tax, 2),
            "paid": round(paid, 2),
            "unpaid": round(unpaid, 2),
        },
        "gst": {
            "cgst": round(cgst_total, 2),
            "sgst": round(sgst_total, 2),
            "igst": round(igst_total, 2),
        },
        "top_customers": [{"name": n, "amount": round(a, 2)} for n, a in top_customers],
        "top_partners": [{"name": n, "amount": round(a, 2)} for n, a in top_partners],
        "monthly": [{"month": m, "amount": round(a, 2)} for m, a in monthly],
    }


# Health endpoint is defined on public_router above (line ~432).

